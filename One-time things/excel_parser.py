import openpyxl
from collections import Counter

workbook = openpyxl.load_workbook('Стата на данный момент.xlsx', data_only=True, read_only=True)

worksheet = workbook.worksheets[0]

row_of_ghosts = list()

for row in range(1, worksheet.max_row + 1):
    value = worksheet.cell(row, 5).value
    if value and value not in ('Сережа', 'Саня'):
        row_of_ghosts.append(value)

count_of_ghosts = Counter(row_of_ghosts)
print(count_of_ghosts)

